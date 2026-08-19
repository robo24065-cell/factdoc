# nk-opinion-parse.py — 서울대 통일평화연구원 「통일의식조사」 원자료 디코더
#
# nk-opinion-harvest.mjs 가 호출하는 포맷 디코더 전용 헬퍼다. 네트워크를 타지 않는다.
# 판단·집계 로직은 전부 mjs 쪽에 있고, 여기서는 "파일 → 셀/좌표" 로만 내려준다.
#
#   python nk-opinion-parse.py xlsx      <path>   → {sheets:[{name, rows:[[cell,…],…]}]}
#   python nk-opinion-parse.py report    <path>   → 기초보고서 PDF 의 「표 2. 남북한 통일의 필요성」
#                                                   전체 행을 좌표 기반으로 추출
#   python nk-opinion-parse.py microdata <path>   → 로데이터 ZIP 내 xlsx 요약 + 가중 집계
#
# 출력은 항상 stdout 에 ensure_ascii JSON 1줄. 실패 시 stderr 로 사유를 쓰고 exit 1.
#
# ── 기초보고서 PDF 가 왜 좌표 기반이어야 하는가 (실측) ──────────────────────────
#   같은 표인데 연도마다 열 순서가 다르다.
#     2021·2022 : ① ② ①+② ③ ④ ⑤ ④+⑤ 계          (8열)
#     2023      : ① ② ③ ④ ⑤ ①+② ④+⑤ 계          (8열)
#     2024·2025 : ① ② ③ ④ ⑤ ①+② ③ ④+⑤ 계       (9열, ③ 이 두 번)
#   게다가 2022·2023 판은 텍스트 추출 순서가 뒤엉켜(“제 장조사 결과 집계표 2 .”) 흐름 기반
#   파싱이 불가능하다. 그래서 열 x좌표로 머리글을 되붙여 ①+② / ③ / ④+⑤ 를 식별한다.
import sys, json, io, re, zipfile

def die(msg):
    sys.stderr.write(msg + "\n")
    sys.exit(1)

def emit(obj):
    sys.stdout.buffer.write(json.dumps(obj, ensure_ascii=True).encode("ascii"))
    sys.stdout.buffer.write(b"\n")

# ── xlsx → 셀 격자 ──────────────────────────────────────────────────────────
def cmd_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for r in ws.iter_rows(values_only=True):
            out = []
            for c in r:
                if c is None:
                    out.append(None)
                elif isinstance(c, (int, float)):
                    out.append(c)
                else:
                    s = str(c).strip()
                    out.append(s if s != "" else None)
            while out and out[-1] is None:
                out.pop()
            rows.append(out)
        while rows and not rows[-1]:
            rows.pop()
        sheets.append({"name": ws.title, "rows": rows})
    emit({"sheets": sheets})

# ── 기초보고서 PDF → 「표 2. 남북한 통일의 필요성」 전체 행 ──────────────────
NUM = re.compile(r"^\d+(?:\.\d+)?$")
SAMPLE = re.compile(r"\((\d[\d,]*)\)")

def _norm(s):
    return re.sub(r"[\s+]", "", s)

def _field_period(win):
    """조사기간 라벨 뒤 창에서 실사기간을 복원. 연도판마다 표기가 다르다(전부 실측):
         2015~2017  '2015년 7월 1일 ∼7월 24일 (24일간)'      끝 연도 생략, 물결표가 ∼
         2018~2020  '조사 기간 3) 년 월 일 월 일 일간 2018 7 12 8 3 (23 )'  각주번호+순서 붕괴
         2021~2025  '2021년 7월 10일 ~ 2021년 8월 4일 (26일간)'
       그래서 구분자에 기대지 않고 숫자 토큰만 뽑은 뒤, 마지막 (N일간) 으로 검산한다.
       검산이 어긋나면 추정하지 않고 None 을 돌려준다."""
    import datetime
    toks = [int(x) for x in re.findall(r"\d+", win)]
    start = next((k for k, v in enumerate(toks) if 2000 <= v <= 2099), None)
    if start is None:
        return None
    y1 = toks[start]
    rest = toks[start + 1:]
    if len(rest) >= 5 and 2000 <= rest[2] <= 2099:
        m1, d1, y2, m2, d2 = rest[0], rest[1], rest[2], rest[3], rest[4]
        dur = rest[5] if len(rest) > 5 else None
    elif len(rest) >= 4:
        m1, d1, m2, d2 = rest[0], rest[1], rest[2], rest[3]
        y2 = y1 if m2 >= m1 else y1 + 1
        dur = rest[4] if len(rest) > 4 else None
    else:
        return None
    try:
        a = datetime.date(y1, m1, d1)
        b = datetime.date(y2, m2, d2)
    except ValueError:
        return None
    span = (b - a).days + 1
    if not (0 < span <= 120):
        return None
    if dur is not None and abs(span - dur) > 2:   # (N일간) 검산
        return None
    return {"from": a.isoformat(), "to": b.isoformat(), "days": span, "printedDays": dur}

def cmd_report(path):
    import fitz
    doc = fitz.open(path)
    chars = sum(len(doc[i].get_text()) for i in range(doc.page_count))
    if chars < 2000:
        emit({"textPdf": False, "pages": doc.page_count, "chars": chars,
              "note": "텍스트 레이어 없음(이미지 PDF) — 추출 불가"})
        return

    field = None
    for i in range(min(doc.page_count, 20)):
        t = re.sub(r"\s+", " ", doc[i].get_text())
        lab = re.search(r"조\s*사\s*기\s*간", t)
        if not lab:
            continue
        field = _field_period(t[lab.end(): lab.end() + 200])
        if field:
            field["page"] = i
            break

    target = None
    for pno in range(doc.page_count):
        flat = re.sub(r"\s+", "", doc[pno].get_text())
        if "필요하다고생각하십니까" not in flat:
            continue
        if "사례수" not in flat:          # 설문지 원문 페이지 배제
            continue
        target = pno
        break
    if target is None:
        die("표2(통일 필요성) 집계 페이지를 찾지 못함: " + path)

    page = doc[target]
    words = page.get_text("words")        # (x0,y0,x1,y1,word,block,line,wordno)
    cands = [w for w in words if SAMPLE.search(w[4])]
    if not cands:
        die("전체 행(사례수 괄호값)을 찾지 못함: %s p%d" % (path, target))
    anchor = min(cands, key=lambda w: (round(w[1], 1), w[0]))
    ymid = (anchor[1] + anchor[3]) / 2
    row = sorted([w for w in words if w[1] <= ymid <= w[3]], key=lambda w: w[0])
    header = [w for w in words if w[3] < anchor[1] - 1]

    sample = int(SAMPLE.search(anchor[4]).group(1).replace(",", ""))
    cols = []
    for w in row:
        tok = w[4]
        if not NUM.fullmatch(tok):
            continue
        if w[0] < anchor[2] - 0.5:        # 사례수 열보다 왼쪽 = 라벨 영역
            continue
        xc = (w[0] + w[2]) / 2
        lab = _norm(" ".join(h[4] for h in header if h[0] - 2 <= xc <= h[2] + 2))
        cols.append({"value": float(tok), "header": lab})
    if not cols:
        die("전체 행에서 수치 열을 찾지 못함: %s p%d" % (path, target))

    def pick(pred, what):
        hit = [c for c in cols if pred(c["header"])]
        if not hit:
            die("'%s' 열을 식별하지 못함: %s p%d / headers=%s"
                % (what, path, target, [c["header"] for c in cols]))
        vals = {c["value"] for c in hit}
        if len(vals) > 1:
            die("'%s' 열이 서로 다른 값으로 중복됨: %s → %s" % (what, path, sorted(vals)))
        return hit[0]["value"]

    need    = pick(lambda h: "①②" in h, "필요(①+②)")
    notNeed = pick(lambda h: "④⑤" in h, "불필요(④+⑤)")
    neutral = pick(lambda h: "반반" in h and "①②" not in h and "④⑤" not in h
                             and "④" not in h and "⑤" not in h, "반반/보통(③)")
    total = need + neutral + notNeed
    if not (99.0 <= total <= 101.0):
        die("필요+반반+불필요 = %.1f (100 아님) — %s p%d" % (total, path, target))

    emit({"textPdf": True, "pages": doc.page_count, "page": target,
          "sampleSize": sample, "need": need, "neutral": neutral, "notNeed": notNeed,
          "sum": round(total, 1), "fieldPeriod": field,
          "columns": [c["header"] for c in cols]})

# ── 로데이터 ZIP → 파일목록 + xlsx 요약 + 가중 집계 ─────────────────────────
def cmd_microdata(path):
    import openpyxl
    z = zipfile.ZipFile(path)
    files = []
    for i in z.infolist():
        raw = i.filename
        try:
            name = raw.encode("cp437").decode("cp949")   # 한글 ZIP 엔트리 복원
        except Exception:
            name = raw
        files.append({"name": name, "bytes": i.file_size})
    xl = [i for i in z.infolist() if i.filename.lower().endswith(".xlsx")]
    if not xl:
        emit({"files": files, "xlsx": None, "note": "ZIP 안에 xlsx 없음"})
        return
    wb = openpyxl.load_workbook(io.BytesIO(z.read(xl[0])), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    kor = list(next(it))
    code = list(next(it))                 # 2행 = 변수명(uni01_a, wt …)
    idx = {c: i for i, c in enumerate(code) if c}
    jq, jw = idx.get("uni01_a"), idx.get("wt")
    n = 0
    raw_cnt, w_cnt, w_sum = {}, {}, 0.0
    for r in it:
        if r[0] is None or str(r[0]).strip() == "":
            continue
        n += 1
        if jq is not None:
            k = r[jq]
            raw_cnt[k] = raw_cnt.get(k, 0) + 1
            if jw is not None and isinstance(r[jw], (int, float)):
                w = float(r[jw])
                w_cnt[k] = w_cnt.get(k, 0.0) + w
                w_sum += w
    weighted = None
    if w_sum > 0:
        pc = {str(k): round(100 * v / w_sum, 1) for k, v in w_cnt.items()}
        g = lambda *ks: round(sum(100 * w_cnt.get(k, 0.0) / w_sum for k in ks), 1)
        weighted = {"variable": "uni01_a", "weight": "wt", "byCode": pc,
                    "need": g(1, 2), "neutral": g(3), "notNeed": g(4, 5)}
    emit({"files": files,
          "xlsx": {"entry": [f["name"] for f in files if f["name"].lower().endswith(".xlsx")][0],
                   "sheet": ws.title, "rows": ws.max_row, "cols": ws.max_column,
                   "headerRows": 2, "respondents": n,
                   "columnsKor": [str(c) for c in kor[:20] if c is not None],
                   "columnsVar": [str(c) for c in code[:20] if c is not None],
                   "rawCount": {str(k): v for k, v in raw_cnt.items()},
                   "weighted": weighted}})

if __name__ == "__main__":
    if len(sys.argv) < 3:
        die("usage: nk-opinion-parse.py <xlsx|report|microdata> <path>")
    cmd, p = sys.argv[1], sys.argv[2]
    if cmd == "xlsx":        cmd_xlsx(p)
    elif cmd == "report":    cmd_report(p)
    elif cmd == "microdata": cmd_microdata(p)
    else:                    die("unknown command: " + cmd)
